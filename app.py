# pip install flask
from flask import Flask, render_template, request
app = Flask(__name__)

# 크롤링 라이브러리 import
import requests
from bs4 import BeautifulSoup

import time
from selenium import webdriver
from selenium.webdriver.common.by import By

# 엑셀 쓰기 위한 준비
from openpyxl import Workbook
write_wb = Workbook()
write_ws = write_wb.active

# write_ws.cell(1, 1, "안녕")
# write_wb.save("result.xlsx")

driver = webdriver.Chrome()

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches',['enable-logging'])
driver = webdriver.Chrome(options=options)
#----------------------------------------------------------------------
@app.route('/')
def hello_world() :
    return render_template("index.html")
#----------------------------------------------------------------------
@app.route('/result', methods=['POST'])
def result():

    if request.method == 'POST' :
        keyword = request.form['input1']
        page = request.form['input2']
        daum_list = []

        for num in range(1, int(page)+1):
            url = "https://search.daum.net/search?w=news&nil_search=btn&DA=NTB&enc=utf8&cluster=y&cluster_page=1&q=" + keyword + "&p=" + str(num)
            req = requests.get(url)
            soup = BeautifulSoup(req.text, "html.parser")

            for i in soup.find_all("a", class_="tit_main fn_tit_u"):
                print(i.text)
                daum_list.append(i.text)
    
        for i in range(1, len(daum_list)+1):
            write_ws.cell(i, 1, daum_list[i-1])
        write_wb.save("static/result.xlsx")

        return render_template("result.html", daum_list = daum_list)
#----------------------------------------------------------------------
@app.route('/naver_shopping', methods = ['POST'])
def naver_shopping() :

    search = request.form['input3']
    search_list = []
    search_list_src = []
    driver = webdriver.Chrome(options=options)
    
    driver.implicitly_wait(3)
    
    driver.get("https://search.shopping.naver.com/search/all?query=" + search)

    #스크롤 내리기
    y=1000
    for timer in range(0,9):
        driver.execute_script("window.scrollTo(0, " + str(y) + ")" )
        y=y+1000
        time.sleep(1)

    soup = BeautifulSoup(driver.page_source, 'html.parser')

    for i in soup.find_all("div", attrs={"class": "adProduct_title__amInq"}):
        # print(i.text)
        search_list.append(i.text)

    for i in soup.find_all("div", attrs={"class": "adProduct_img_area__wPZ_E"}):
        # print(i.find("img")['src'])              
        search_list_src.append(i.find("img")['src'])

    for i in soup.find_all("div", attrs={"class": "product_title__Mmw2K"}):
        # print(i.text)
        search_list.append(i.text)

    for i in soup.find_all("div", attrs={"class": "product_img_area__cUrko"}):
        # print(i.find("img")['src'])              
        search_list_src.append(i.find("img")['src'])

    print("-------0-0-0-0-0-0-0-------")
    # 네이버 쇼핑에서 해외직구 버튼을 눌러서 이동

    driver.find_element(By.CSS_SELECTOR, "#content > div.style_content__xWg5l > div.seller_filter_area > ul > li:nth-child(6) > a").click()
    time.sleep(1)
    #스크롤 내리기
    y=1000
    for timer in range(0,9):
        driver.execute_script("window.scrollTo(0, " + str(y) + ")" )
        y=y+1000
        time.sleep(1)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    for i in soup.find_all("a", attrs={"class": "product_link__TrAac linkAnchor"}):
        print(i.text)
        search_list.append(i.text)

    #img
    for i in soup.find_all("a", attrs={"class": "thumbnail_thumb__Bxb6Z linkAnchor"}):
        print(i.find("img")['src']) 
        search_list_src.append(i.find("img")['src'])

    driver.close()

    return render_template("shopping.html", 
                                search_list=search_list,
                                search_list_src = search_list_src,
                                len = len(search_list_src))
#----------------------------------------------------------------------
if __name__ == "__main__":
    app.run()